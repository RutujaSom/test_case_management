import frappe
import csv
import io
from openpyxl import load_workbook
import pandas as pd 


@frappe.whitelist()
def get_test_cases_query_for_project(doctype, txt, searchfield, start, page_len, filters):
    module = filters.get('custom_module') or ''
    
    return frappe.db.sql("""
        SELECT
            name,
            test_case_id,
            title
        FROM `tabTest Case Bank`
        WHERE
            (test_case_id LIKE %(txt)s OR title LIKE %(txt)s)
            AND module LIKE %(module)s
        ORDER BY creation DESC
        LIMIT %(start)s, %(page_len)s
    """, {
        "txt": f"%{txt}%",
        "module": f"%{module}%",
        "start": start,
        "page_len": page_len
    }, as_dict=True)


@frappe.whitelist()
def import_test_cases_from_file_for_bank(file_url, file_type):
    """
    Import test cases from uploaded CSV or Excel file into Test Case Bank
    """
    file_doc = frappe.get_doc("File", {"file_url": file_url})
    file_content = file_doc.get_content()
    filename = file_doc.file_name.lower()

    # validate extension
    if file_type == "CSV" and not filename.endswith(".csv"):
        frappe.throw("Selected file type is CSV but uploaded file is not a .csv file.")
    elif file_type == "Excel" and not filename.endswith(".xlsx"):
        frappe.throw("Selected file type is Excel but uploaded file is not a .xlsx file.")

    rows = []
    if filename.endswith(".xlsx"):
        file_path = file_doc.get_full_path()
        with open(file_path, "rb") as f:
            wb = load_workbook(f, data_only=True)
            ws = wb.active
            headers = [cell.value for cell in ws[1]]
            for row in ws.iter_rows(min_row=2, values_only=True):
                rows.append(dict(zip(headers, row)))
    elif filename.endswith(".csv"):
        if isinstance(file_content, bytes):
            try:
                decoded_content = file_content.decode("utf-8")
            except UnicodeDecodeError:
                decoded_content = file_content.decode("latin-1")
        else:
            decoded_content = file_content
        df = pd.read_csv(io.StringIO(decoded_content), keep_default_na=None)
        rows = df.to_dict(orient="records")

    # --- Status normalization map ---
    STATUS_MAP = {
        "open": "Draft",
        "draft": "Draft",
        "approved": "Approved",
        "obsolete": "Obsolete",
        "closed": "Obsolete"
    }
    VALID_STATUSES = ["Draft", "Approved", "Obsolete"]

    count = 0
    skipped = []

    for idx, row in enumerate(rows, start=1):
        try:
            title = (row.get("test_case") or "").strip()
            description = (row.get("description") or "").strip()
            expected_result = (row.get("expected_result") or "").strip()
            pre_conditions = (row.get("pre_conditions") or "").strip()
            test_case_id = (row.get("test_case_id") or "").__str__()
            priority = (row.get('priority') or "").strip()
            type = (row.get('type') or "").strip()
            module = (row.get("module") or "").strip()
            raw_steps = (row.get("steps") or "").strip()

            # --- normalize status ---
            raw_status = (row.get("status") or "Draft").strip().lower()
            status = STATUS_MAP.get(raw_status, "Draft")
            if status not in VALID_STATUSES:
                status = "Draft"

            # --- validation ---
            if not type:
                skipped.append(f"Row {idx}: Missing type")
                continue
            type_doc_name = frappe.db.get_value("Test Case Type", {"title": type})
            if not type_doc_name:
                skipped.append(f"Row {idx}: Invalid type '{type}'")
                continue

            # --- handle module gracefully ---
            module_def_name = None
            if module:
                module_def_name = frappe.db.get_value("Test Case Module", {"module_name": module})
                if not module_def_name:
                    skipped.append(f"Row {idx}: Module '{module}' not found → Test Case created without module")

            if not title or not expected_result:
                skipped.append(f"Row {idx}: Missing title/expected_result")
                continue

            # --- create test case ---
            test_case = frappe.new_doc("Test Case Bank")
            test_case.title = title
            test_case.test_case_id = test_case_id
            test_case.description = description
            test_case.pre_conditions = pre_conditions
            test_case.expected_results = expected_result
            test_case.status = status
            test_case.priority = priority
            test_case.test_case_type = type_doc_name
            test_case.custom_module = module_def_name

            test_case.append("case_steps", {
                "title": raw_steps,
                "step_completed": 0
            })

            test_case.insert(ignore_permissions=True)
            count += 1

        except Exception as e:
            skipped.append(f"Row {idx}: Error {str(e)}")
            frappe.log_error(
                message=frappe.get_traceback(),
                title="Import Test Case Error"
            )

    frappe.db.commit()

    return f"""
✅ {count} Test Case(s) imported into Test Case Bank.  
⚠️ Skipped Rows: {len(skipped)}  
{chr(10).join(skipped)}
    """

