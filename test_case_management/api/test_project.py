

import frappe
import csv
import io
from openpyxl import load_workbook
import pandas as pd


@frappe.whitelist()
def import_test_cases_from_file_for_project(file_url, project, file_type):
    """
    Import Test Cases from CSV/Excel into the Test Case DocType,
    and link them to a given Test Project.
    Only assigns a module if it already exists (does NOT auto-create new modules).
    """
    file_doc = frappe.get_doc("File", {"file_url": file_url})
    file_content = file_doc.get_content()
    filename = file_doc.file_name.lower()

    # --- Validate extension ---
    if file_type == "CSV" and not filename.endswith(".csv"):
        frappe.throw("Selected file type is CSV but uploaded file is not a .csv file.")
    elif file_type == "Excel" and not filename.endswith(".xlsx"):
        frappe.throw("Selected file type is Excel but uploaded file is not a .xlsx file.")

    test_project = frappe.get_doc("Test Project", project)
    rows = []

    # --- Read Excel file ---
    if filename.endswith(".xlsx"):
        file_path = file_doc.get_full_path()
        with open(file_path, "rb") as f:
            wb = load_workbook(f, data_only=True)
            ws = wb.active
            headers = [cell.value for cell in ws[1]]
            for row in ws.iter_rows(min_row=2, values_only=True):
                rows.append(dict(zip(headers, row)))

    # --- Read CSV file ---
    elif filename.endswith(".csv"):
        if isinstance(file_content, bytes):
            try:
                decoded_content = file_content.decode("utf-8")
            except UnicodeDecodeError:
                decoded_content = file_content.decode("latin-1")
        else:
            decoded_content = file_content
        df = pd.read_csv(io.StringIO(decoded_content), keep_default_na=None)
        rows = df.to_dict(orient="records")

    # --- Status normalization map ---
    STATUS_MAP = {
        "open": "Draft",
        "draft": "Draft",
        "approved": "Approved",
        "obsolete": "Obsolete",
        "closed": "Obsolete"
    }
    VALID_STATUSES = ["Draft", "Approved", "Obsolete"]

    count = 0
    skipped = []

    # --- Process rows ---
    for idx, row in enumerate(rows, start=1):
        try:
            title = (row.get("test_case") or "").strip()
            description = (row.get("description") or "").strip()
            expected_result = (row.get("expected_result") or "").strip()
            raw_status = (row.get("status") or "Draft").strip().lower()
            pre_conditions = (row.get("pre_conditions") or "").strip()
            test_case_id = (row.get("test_case_id") or "").__str__()
            priority = (row.get("priority") or "").strip()
            type_name = (row.get("type") or "").strip()
            module_name = (row.get("module") or "").strip()
            raw_steps = (row.get("steps") or "").strip()

            # --- normalize status ---
            status = STATUS_MAP.get(raw_status, "Draft")
            if status not in VALID_STATUSES:
                status = "Draft"

            # --- validate type ---
            if not type_name:
                skipped.append(f"Row {idx}: Missing type")
                continue
            type_doc_name = frappe.db.get_value("Test Case Type", {"title": type_name})
            if not type_doc_name:
                # auto-create type if missing
                new_type = frappe.get_doc({
                    "doctype": "Test Case Type",
                    "title": type_name
                })
                new_type.insert(ignore_permissions=True)
                type_doc_name = new_type.name

            # --- handle module only if it exists ---
            module_value = None
            if module_name:
                module_field = frappe.get_meta("Test Case").get_field("custom_module")
                if module_field and module_field.fieldtype == "Link" and module_field.options == "Test Case Module":
                    # only link if module exists
                    module_def_name = frappe.db.get_value("Test Case Module", {"module_name": module_name})
                    if module_def_name:
                        module_value = module_def_name
                else:
                    # if field is Data, assign value directly
                    module_value = module_name

            # --- validate required fields ---
            if not title or not expected_result:
                skipped.append(f"Row {idx}: Missing title/expected_result")
                continue

            # --- create test case ---
            test_case = frappe.new_doc("Test Case")
            test_case.title = title
            test_case.test_case_id = test_case_id
            test_case.description = description
            test_case.pre_conditions = pre_conditions
            test_case.expected_results = expected_result
            test_case.project = project
            test_case.status = status
            test_case.priority = priority
            test_case.test_case_type = type_doc_name
            test_case.custom_module = module_value  # assign only if valid

            # Add steps
            if raw_steps:
                test_case.append("case_steps", {
                    "title": raw_steps,
                    "step_completed": 0
                })

            test_case.insert(ignore_permissions=True)
            count += 1

        except Exception as e:
            skipped.append(f"Row {idx}: Error {str(e)}")
            frappe.log_error(
                message=frappe.get_traceback(),
                title="Import Test Case Error (Project)"
            )

    frappe.db.commit()

    return f"""
✅ {count} Test Case(s) imported for project {test_project.title}.  
⚠️ Skipped Rows: {len(skipped)}  
{chr(10).join(skipped)}
    """

