import frappe
import csv
import io
from openpyxl import load_workbook
import pandas as pd 


@frappe.whitelist()
def import_test_cases_from_file(file_url, test_plan, file_type):
    """
    Import test cases from an uploaded file (CSV or Excel) and link them to a Test Plan.
    """

    # Get the file document and content
    file_doc = frappe.get_doc("File", {"file_url": file_url})
    file_content = file_doc.get_content()
    filename = file_doc.file_name.lower()

    # Validate uploaded file type against selected file_type
    if file_type == "CSV" and not filename.endswith(".csv"):
        frappe.throw("Selected file type is CSV but uploaded file is not a .csv file.")
    elif file_type == "Excel" and not filename.endswith(".xlsx"):
        frappe.throw("Selected file type is Excel but uploaded file is not a .xlsx file.")

    # Fetch the Test Plan and its associated project
    test_plan_doc = frappe.get_doc("Test Plan", test_plan)
    project = test_plan_doc.project
    count = 0
    rows = []

    # Read Excel file and convert rows to dictionary
    if filename.endswith(".xlsx"):
        file_path = file_doc.get_full_path()
        with open(file_path, "rb") as f:
            wb = load_workbook(f, data_only=True)
            ws = wb.active
            headers = [cell.value for cell in ws[1]]  # First row as headers
            for row in ws.iter_rows(min_row=2, values_only=True):
                rows.append(dict(zip(headers, row)))

    # Read CSV file and convert rows to dictionary
    elif filename.endswith(".csv"):
        if isinstance(file_content, bytes):
            try:
                decoded_content = file_content.decode("utf-8")
            except UnicodeDecodeError:
                decoded_content = file_content.decode("latin-1")
        else:
            decoded_content = file_content

        df = pd.read_csv(io.StringIO(decoded_content), keep_default_na=None)
        rows = df.to_dict(orient="records")

    # Process each row to create Test Case
    for row in rows:
        title = (row.get("test_case") or "").strip()
        description = (row.get("description") or "").strip()
        expected_result = (row.get("expected_result") or "").strip()
        status = (row.get("status") or "Open").strip()
        pre_conditions = (row.get("pre_conditions") or "").strip()
        test_case_id = (row.get("test_case_id") or "").strip()
        priority = (row.get('priority') or "").strip()
        type = (row.get('type') or "").strip()

        # Skip if required 'type' is missing
        if not type:
            print(f"Skipping row due to missing type (required): {row}")
            continue

        # Find the Test Case Type using title field
        type_doc_name = frappe.db.get_value("Test Case Type", {"title": type})
        if not type_doc_name:
            print(f"Skipping row: Invalid type '{type}' not found in Test Case Type.")
            continue

        # Steps can be a string like: "Step 1||Step 2"
        raw_steps = (row.get("steps") or "").strip()

        # Skip if title or expected result is missing
        if not title or not expected_result:
            print(f"Skipping row due to missing title or expected_result: {row}")
            continue

        # Create new Test Case
        test_case = frappe.new_doc("Test Case")
        test_case.title = title
        test_case.test_case_id = test_case_id
        test_case.description = description
        test_case.pre_conditions = pre_conditions
        test_case.expected_results = expected_result
        test_case.project = project
        test_case.status = status
        test_case.priority = priority
        test_case.test_case_type = type_doc_name

        # Add steps to child table
        test_case.append("case_steps", {
            "title": raw_steps,
            "step_completed": 0  # default to not completed
        })

        # Insert Test Case into database
        test_case.insert(ignore_permissions=True)

        # Link to Test Plan
        test_plan_doc.append("test_cases", {
            "test_case": test_case.name,
            "test_case_title": test_case.title
        })
        count += 1

    # Save and commit the updated Test Plan
    test_plan_doc.save()
    frappe.db.commit()

    return f"{count} Test Case(s) imported and linked to Test Plan."


# @frappe.whitelist()
# def download_template(file_type="Excel"):
#     """
#     Provide a blank template (CSV or Excel) with required headers for importing test cases.
#     """

#     headers = [
#         "test_case_id", "test_case", "steps",
#         "pre_conditions", "expected_result",
#         "priority", "type", "module"
#     ]
    
#     df = pd.DataFrame(columns=headers)

#     if file_type == "CSV":
#         # Generate CSV file content
#         output = io.StringIO()
#         df.to_csv(output, index=False)
#         content = output.getvalue()

#         # Set response metadata for file download
#         frappe.response.filename = "test_case_template.csv"
#         frappe.response.filecontent = content
#         frappe.response.type = "binary"
#     else:
#         # Generate Excel file content
#         output = io.BytesIO()
#         with pd.ExcelWriter(output, engine="openpyxl") as writer:
#             df.to_excel(writer, index=False)

#         # Set response metadata for file download
#         frappe.response.filename = "test_case_template.xlsx"
#         frappe.response.filecontent = output.getvalue()
#         frappe.response.type = "binary"



@frappe.whitelist()
def download_template(file_type="Excel"):
    """
    Provide a blank template (CSV or Excel) with required headers for importing test cases.
    Includes default priority values: High, Medium, Low
    """

    headers = [
        "test_case_id", "test_case", "steps",
        "pre_conditions", "expected_result",
        "priority", "type", "module"
    ]

    # Add default priority values
    sample_data = [
        {"test_case_id": "", "test_case": "", "steps": "", "pre_conditions": "",
         "expected_result": "", "priority": "High", "type": "", "module": ""},
        {"test_case_id": "", "test_case": "", "steps": "", "pre_conditions": "",
         "expected_result": "", "priority": "Medium", "type": "", "module": ""},
        {"test_case_id": "", "test_case": "", "steps": "", "pre_conditions": "",
         "expected_result": "", "priority": "Low", "type": "", "module": ""}
    ]

    df = pd.DataFrame(sample_data, columns=headers)

    if file_type == "CSV":
        output = io.StringIO()
        df.to_csv(output, index=False)
        content = output.getvalue()

        frappe.response.filename = "test_case_template.csv"
        frappe.response.filecontent = content
        frappe.response.type = "binary"
    else:
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            df.to_excel(writer, index=False)

        frappe.response.filename = "test_case_template.xlsx"
        frappe.response.filecontent = output.getvalue()
        frappe.response.type = "binary"
